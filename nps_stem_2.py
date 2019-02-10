import xlrd
import pysolr as solr
import xlsxwriter
import datetime
import os
import json
import tensorflow as tf
from tensorflow.contrib import learn
import data_helper
import numpy as np
import logging
from urllib.request import urlopen
import language_process as lp
import openpyxl
import pyexcelerate
from datetime import datetime


def nps_data_pull():
    current_directory = os.getcwd()
    os.chdir(solr_path)
    os.system('solr delete -c ' + table_name)
    os.system('solr create -c ' + table_name)
    os.chdir(current_directory)

    reader = xlrd.open_workbook(file_path + 'Scotiabank_Responses_Sept.xlsx')
    #reader = xlrd.open_workbook(file_path + 'sentiment_model_test_data.xlsx')
    sh = reader.sheet_by_index(0)
    s = solr.Solr(url_path + table_name)
    i=1
    doc=[]
    for rownum in range(3, sh.nrows):
        if len(sh.row_values(rownum)[7])>= 6 or len(sh.row_values(rownum)[8]) >= 6:
         #   why_stem_str=" ".join(lp.tokenize_lower(str(sh.row_values(rownum)[7])))
           # what_stem_str=" ".join(lp.tokenize_lower(str(sh.row_values(rownum)[8])))
            #print(str(sh.row_values(rownum)[6]))
            date=str(sh.row_values(rownum)[5]).split('/')[0]
            doc.append(dict(
                publish_date=str(sh.row_values(rownum)[5]),
                first_name=str(sh.row_values(rownum)[0]),
                last_name=str(sh.row_values(rownum)[1]),
                activity=str(sh.row_values(rownum)[2]),
                channel_code=str(sh.row_values(rownum)[3]),
                customer_id=str(sh.row_values(rownum)[4]),
                response_date=str(sh.row_values(rownum)[5]),
                rating=str(sh.row_values(rownum)[6]),
                why=str(sh.row_values(rownum)[7]),
                what=str(sh.row_values(rownum)[8]),
                system_date=str(date),
                sentiment='yes',
                non_digital='no',
                flag_why=0,
                flag_what=0,
                category='init',
                id_val=i
            ))
            i=i+1
        else:
            date=str(sh.row_values(rownum)[5]).split('/')[0]
            void.append(dict(
                publish_date=str(sh.row_values(rownum)[5]),
                first_name=str(sh.row_values(rownum)[0]),
                last_name=str(sh.row_values(rownum)[1]),
                activity=str(sh.row_values(rownum)[2]),
                channel_code=str(sh.row_values(rownum)[3]),
                customer_id=str(sh.row_values(rownum)[4]),
                response_date=str(sh.row_values(rownum)[5]),
                rating=str(sh.row_values(rownum)[6]),
                why=str(sh.row_values(rownum)[7]),
                what=str(sh.row_values(rownum)[8]),
                system_date=str(date),
                sentiment='yes',
                non_digital='no',
                flag_why=0,
                flag_what=0,
                category='Void_Short_Comment',
                id_val=i
            ))
            i=i+1
            
    s.add(doc)
    s.commit()

def nps_prediction():
    print("Sentiment being processed")

    publish_date=[]
    first_name=[]
    last_name=[]
    activity=[]
    channel_code=[]
    customer_id=[]
    response_date=[]
    rating=[]
    why=[]
    what=[]
    system_date=[]
    sentiment_why=[]
    sentiment_what = []
    sentiment=[]
    non_digital_why=[]
    non_digital_what=[]
    non_digital=[]
    id_val=[]

    s = solr.Solr('http://localhost:8983/solr/' + table_name)

    url_call = 'http://localhost:8983/solr/' + table_name + '/select?indent=on&q=*:*&wt=python'
    conn = urlopen(url_call)
    resp = eval(conn.read())
    resp_rows = str(resp['response']['numFound'])
    print(resp_rows) #How many record
    url_string = url_call + '&rows=' + resp_rows + '&start=0'
    print(url_string)
    conn = urlopen(url_string)
    resp = eval(conn.read())

    for doc1 in resp['response']['docs']:
        if 'why' in doc1.keys():
            why.append(doc1['why'][0])
        else:
            why.append('no comment')
        if 'what' in doc1.keys():
            what.append(doc1['what'][0])
        else:
            what.append('no comment')
        if 'last_name' in doc1.keys():
            last_name.append(doc1['last_name'][0])
        else:
            last_name.append('Business User')
        publish_date.append(doc1['publish_date'][0])
        first_name.append(doc1['first_name'][0])
        activity.append(doc1['activity'][0])
        channel_code.append(doc1['channel_code'][0])
        customer_id.append(doc1['customer_id'][0])
        response_date.append(doc1['response_date'][0])
        rating.append(doc1['rating'][0])
        system_date.append(doc1['system_date'][0])
        #Initial Flags
        sentiment.append('no')
        non_digital.append('no')
        id_val.append(doc1['id_val'][0])

    for prediction in ['sentiment','digital']:
        if prediction == 'sentiment':
            params = json.loads(open('./parameters_sentiment.json').read())
            checkpoint_dir = "C:\\Code_Sketch\\NPS\\S3134076\\PycharmProjects\\nps\\trained_model_1509637563"
            labels = json.loads(open('./labels_sentiment.json').read())
            vocab_path = os.path.join(checkpoint_dir, "vocab_sentiment.pickle")
        elif prediction == 'digital':
            params = json.loads(open('./parameters_digital.json').read())
            checkpoint_dir = "C:\\Code_Sketch\\NPS\\S3134076\\PycharmProjects\\nps\\trained_model_1508867627"
            labels = json.loads(open('./labels_digital.json').read())
            vocab_path = os.path.join(checkpoint_dir, "vocab_digital.pickle")

        logging.getLogger().setLevel(logging.INFO)

       # if not checkpoint_dir.endswith('/'):
        #    checkpoint_dir += '/'
        checkpoint_file = tf.train.latest_checkpoint(checkpoint_dir + '\\checkpoints')
        logging.critical('Loaded the trained model: {}'.format(checkpoint_file))

        one_hot = np.zeros((len(labels), len(labels)), int)
        np.fill_diagonal(one_hot, 1)



        vocab_processor = learn.preprocessing.VocabularyProcessor.restore(vocab_path)
        x_test_why = [data_helper.clean_str(x) for x in why]
        x_test_why = np.array(list(vocab_processor.transform(x_test_why)))

        x_text_what=[data_helper.clean_str(x) for x in what]
        x_test_what = np.array(list(vocab_processor.transform(x_text_what)))

        graph = tf.Graph()
        with graph.as_default():
            session_conf = tf.ConfigProto(allow_soft_placement=True, log_device_placement=False)
            sess = tf.Session(config=session_conf)

            with sess.as_default():
                saver = tf.train.import_meta_graph("{}.meta".format(checkpoint_file))
                saver.restore(sess, checkpoint_file)

                input_x_why = graph.get_operation_by_name("input_x").outputs[0]
                dropout_keep_prob_why = graph.get_operation_by_name("dropout_keep_prob").outputs[0]
                predictions_why = graph.get_operation_by_name("output/predictions").outputs[0]

                input_x_what = graph.get_operation_by_name("input_x").outputs[0]
                dropout_keep_prob_what = graph.get_operation_by_name("dropout_keep_prob").outputs[0]
                predictions_what = graph.get_operation_by_name("output/predictions").outputs[0]


                batches_why = data_helper.batch_iter(list(x_test_why), params['batch_size'], 1, shuffle=False)
                batches_what= data_helper.batch_iter(list(x_test_what), params['batch_size'], 1, shuffle=False)

                all_predictions_why = []
                all_predictions_what = []

                for x_test_batch in batches_why:
                    batch_predictions_why = sess.run(predictions_why, {input_x_why: x_test_batch, dropout_keep_prob_why: 1.0})
                    all_predictions_why = np.concatenate([all_predictions_why, batch_predictions_why])
                # Giving Prediction batch by batch

                for x_test_batch in batches_what:
                    batch_predictions_what = sess.run(predictions_what, {input_x_what: x_test_batch, dropout_keep_prob_what: 1.0})
                    all_predictions_what = np.concatenate([all_predictions_what, batch_predictions_what])



        for predict in all_predictions_why:
            #print(labels[int(predict)])
            if prediction == 'sentiment':
                if labels[int(predict)] == 'positive':
                    sentiment_why.append("positive")
                else:
                    sentiment_why.append("negative")
            elif prediction == 'digital':
                if labels[int(predict)] == 'non_digital':
                    non_digital_why.append("non_digital")
                else:
                    non_digital_why.append("digital")
        for predict in all_predictions_what:
            if prediction == 'sentiment':
                if labels[int(predict)] == 'positive':
                    sentiment_what.append("positive")
                else:
                    sentiment_what.append("negative")
            elif prediction == 'digital':
                if labels[int(predict)] == 'non_digital':
                    non_digital_what.append("non_digital")
                else:
                    non_digital_what.append("digital")


    current_directory = os.getcwd()
    os.chdir(solr_path)
    os.system('solr delete -c ' + table_name)
    os.system('solr create -c ' + table_name)
    os.chdir(current_directory)

    doc=[]
   # x=2
   """puting to D/ND,P/N"""
    for i in range(3, len(channel_code)):
        if (sentiment_why[i]=='negative' or sentiment_what[i]=='negative') and \
        (non_digital_why[i]=='digital' or non_digital_what[i]=='digital'):
            why_stem_str=" ".join(lp.tokenize_lower(str(why[i])))
            what_stem_str=" ".join(lp.tokenize_lower(str(what[i])))
            doc.append(dict(
                publish_date=publish_date[i],
                first_name=first_name[i],
                last_name=last_name[i],
                activity=activity[i],
                channel_code=channel_code[i],
                customer_id=customer_id[i],
                response_date=response_date[i],
                rating=rating[i],
                why=why[i],
                what=what[i],
               # what_stem=lp.tokenize_lower(what[i]),
                system_date=system_date[i],
                sentiment_why=sentiment_why[i],
                sentiment_what=sentiment_what[i],
                non_digital_why=non_digital_why[i],
                non_digital_what=non_digital_what[i],
                id_val=id_val[i],
                flag_why=0,
                flag_what=0,
                category=nps_categorization(why_stem_str+' '+what_stem_str,0)
            ))
            print("upload"+str(i))
            #if "catagory" not in doc.keys():
                
        elif (sentiment_why[i]=='negative' or sentiment_what[i]=='negative') and \
        (non_digital_why[i]=='non_digital' and non_digital_what[i]=='non_digital'):
            nond_negative.append(dict(
                publish_date=publish_date[i],
                first_name=first_name[i],
                last_name=last_name[i],
                activity=activity[i],
                channel_code=channel_code[i],
                customer_id=customer_id[i],
                response_date=response_date[i],
                rating=rating[i],
                why=why[i],
                what=what[i],
                system_date=system_date[i],
                sentiment_why=sentiment_why[i],
                sentiment_what=sentiment_what[i],
                non_digital_why=non_digital_why[i],
                non_digital_what=non_digital_what[i],
                id_val=id_val[i],
                flag_why=0,
                flag_what=0,
                category=nps_categorization(why[i]+' '+what[i],1)
            ))
        elif (sentiment_why[i]=='positive' and  sentiment_what[i]=='positive') and\
        (non_digital_why[i]=='digital' or non_digital_what[i]=='digital'):
            digital_positive.append(dict(
                publish_date=publish_date[i],
                first_name=first_name[i],
                last_name=last_name[i],
                activity=activity[i],
                channel_code=channel_code[i],
                customer_id=customer_id[i],
                response_date=response_date[i],
                rating=rating[i],
                why=why[i],
                what=what[i],
                system_date=system_date[i],
                sentiment_why=sentiment_why[i],
                sentiment_what=sentiment_what[i],
                non_digital_why=non_digital_why[i],
                non_digital_what=non_digital_what[i],
                id_val=id_val[i],
                flag_why=0,
                flag_what=0,
                category=["Digital_Positive"]
            ))
        elif (sentiment_why[i]=='positive' and  sentiment_what[i]=='positive') and\
        (non_digital_why[i]=='non_digital' and non_digital_what[i]=='non_digital'):
                non_digital_positive.append(dict(
                publish_date=publish_date[i],
                first_name=first_name[i],
                last_name=last_name[i],
                activity=activity[i],
                channel_code=channel_code[i],
                customer_id=customer_id[i],
                response_date=response_date[i],
                rating=rating[i],
                why=why[i],
                what=what[i],
                system_date=system_date[i],
                sentiment_why=sentiment_why[i],
                sentiment_what=sentiment_what[i],
                non_digital_why=non_digital_why[i],
                non_digital_what=non_digital_what[i],
                id_val=id_val[i],
                flag_why=0,
                flag_what=0,
                category=["Non_Digital_Positive"]
            ))   
    
    s.add(doc+nond_negative+digital_positive+non_digital_positive)
    stat['Very_short_Comment']=len(void)
    #stat['Non_Digital_Negative']=len(nond_negative)
    stat["Digital_Positive"]=len(digital_positive)
    stat["Non_Digital_Positive"]=len(non_digital_positive)
    s.commit()

def nps_categorization(str_input,flag):
    doc_input=dict(text=str_input)
    s1 = solr.Solr('http://localhost:8983/solr/' + 'key_word')
    s1.delete(q='*:*')
    s1.commit()
    s1.add([doc_input])
    s1.commit()
    print("categorizing reviews uploading finished")
    if flag == 0:
        """0 is for digital Negative"""
        params_kw = json.loads(open('C:\\Code_Sketch\\NPS\\S3134076\\PycharmProjects\\nps\\Key_word.json').read())
        cata_list=params_kw['cata_list']
        Num_Cata=len(cata_list)
    else:
        params_kw = json.loads(open('C:\\Code_Sketch\\NPS\\S3134076\\PycharmProjects\\nps\\Nond_Key_Word.json').read())
        cata_list=params_kw['cata_list']
        Num_Cata=len(cata_list)
    return_cata_lst=[]
    return_cata_lst_1=[]
    for m in range(0,Num_Cata):
        keywords_list_extract=params_kw[cata_list[m]]
        search_str=''
        for keyword in keywords_list_extract:
            if len(keyword.split(' ')) > 1:
                keyword_str=''
                for key in keyword.split(' '):
                    keyword_str=keyword_str+key+'%20'
                keyword_str=keyword_str[:-3]
            else:
                keyword_str=keyword
            search_str=search_str+'("'+'*'+keyword_str+'*'+'")%20OR%20'
        search_str=search_str[:-8]
        print(search_str)
        url_call = 'http://localhost:8983/solr/' + 'key_word' + '/select?indent=on&q=('+search_str+')&wt=python'
        print(url_call)
        conn = urlopen(url_call)
        resp2 = eval(conn.read())
        #RESP is a dictionary
        resp_rows = str(resp2['response']['numFound'])
        if (int(resp_rows)!=0 and flag==0):
            return_cata_lst.append(cata_list[m])
            stat[cata_list[m]] +=1
        elif (int(resp_rows)!=0 and flag==1):
            print(1111)
            return_cata_lst_1.append(cata_list[m])
            print(return_cata_lst_1)
            stat[cata_list[m]] +=1
    print(return_cata_lst_1)
    if (return_cata_lst==[] and flag==0):
        return_cata_lst=['Navigation']
        stat['Navigation'] +=1
        return return_cata_lst
    elif (return_cata_lst_1==[] and flag==1):
        return_cata_lst_1=['Non_Digital_Others']
        stat['Non_Digital_Others'] +=1
        return return_cata_lst_1
    elif (return_cata_lst!=[] and flag==0):
        return return_cata_lst
    elif (return_cata_lst_1!=[] and flag==1):
        return return_cata_lst_1
    
            
    

def data_writer():
   # workbook_n = openpyxl.load_workbook(file_path + "Data_writer_Result(2).xlsx")
    workbook_ne = xlsxwriter.Workbook(file_path + "Data_writer_Result.xlsx")
    worksheet_ne = workbook_ne.add_worksheet("result")
    
    params_kw = json.loads(open('C:\\Code_Sketch\\NPS\\S3134076\\PycharmProjects\\nps\\Key_word.json').read())
    params_kw_1 = json.loads(open('C:\\Code_Sketch\\NPS\\S3134076\\PycharmProjects\\nps\\Nond_Key_Word.json').read())
    cata_list_1=["Non_Digital_Positive","Digital_Positive",'Navigation','Non_Digital_Others']+params_kw['cata_list']+params_kw_1['cata_list']
    
        #workbook_n.create_sheet(cata_list[counter+1])
   # worksheet_n=workbook_n.worksheets[0]
    """Change sheet title: worksheet_n.title=cata_list_1[counter]"""
    #row_count = worksheet_n.max_row
    worksheet_ne.write(0, 0, "First Name")
    worksheet_ne.write(0,1,"Last Name")
    worksheet_ne.write(0,2,"Activity Description")
    worksheet_ne.write(0,3,"Channel Code")
    worksheet_ne.write(0,4,"Channel ID")
    worksheet_ne.write(0,5, "Response Date(EST)")
    worksheet_ne.write(0,6,"Likelihood to reccommend")
    worksheet_ne.write(0,7,"Why?")
    worksheet_ne.write(0,8,"What could we do to serve you better?")
    worksheet_ne.write(0,9,"Predicted_Category")
    worksheet_ne.write(0,10,"Sentiment")
    worksheet_ne.write(0,11,"D/ND")
    row_count=1
    #workbook_n.save(file_path + "Data_writer_Result(2).xlsx")
    final=0
    for counter in range(len(cata_list_1)): 
        url_call = 'http://localhost:8983/solr/' + table_name + '/select?q=(category:'+cata_list_1[counter]+')&wt=python'
        print(url_call)
        conn = urlopen(url_call)
        resp = eval(conn.read())
        resp_rows = str(resp['response']['numFound'])
        print(resp_rows)
        url_string = url_call + '&rows=' + resp_rows + '&start=0'
        conn = urlopen(url_string)
        resp = eval(conn.read())
        for doc10 in resp['response']['docs']:
            print("Processing"+str(final))
            worksheet_ne.write(1+final, 0,doc10["first_name"][0])
            worksheet_ne.write(1+final, 1,doc10["last_name"][0])
            worksheet_ne.write(1+final, 2,doc10["activity"][0])
            worksheet_ne.write(1+final, 3,doc10["channel_code"][0])
            worksheet_ne.write(1+final, 4,doc10["customer_id"][0])         
            worksheet_ne.write(1+final, 5,doc10["response_date"][0])
            worksheet_ne.write(1+final, 6,doc10["rating"][0])
            worksheet_ne.write(1+final, 7,doc10["why"][0])
            worksheet_ne.write(1+final, 8,doc10["what"][0])
            worksheet_ne.write(1+final, 9,cata_list_1[counter])
            worksheet_ne.write(1+final, 10,sent_sum(doc10["sentiment_why"][0],doc10["sentiment_what"][0]))
            worksheet_ne.write(1+final, 11,sent_sum(doc10["non_digital_why"][0],doc10["non_digital_what"][0]))
           # workbook_n.save(file_path + "Data_writer_Result(2).xlsx")
            #row_count = worksheet_n.max_row
            final +=1
    for eliment in void:
        print("Processing"+str(final))
        worksheet_ne.write(1+final, 0,eliment["first_name"])
        worksheet_ne.write(1+final, 1,eliment["last_name"])
        worksheet_ne.write(1+final, 2,eliment["activity"])
        worksheet_ne.write(1+final, 3,eliment["channel_code"])
        worksheet_ne.write(1+final, 4,eliment["customer_id"])         
        worksheet_ne.write(1+final, 5,eliment["response_date"])
        worksheet_ne.write(1+final, 6,eliment["rating"])
        worksheet_ne.write(1+final, 7,eliment["why"])
        worksheet_ne.write(1+final, 8,eliment["what"])
        worksheet_ne.write(1+final, 9,eliment['category'])
        worksheet_ne.write(1+final, 10,"Very_Short_Comment")
        worksheet_ne.write(1+final, 11,"Null")
        final +=1
    workbook_ne.close()
    
    
def sent_sum(s1,s2):
    if (s1=="negative" or s2=="negative"):
        return ("negative")
    elif (s1=="digital" or s2=="digital"):
        return ("digital")
    elif (s1=="positive" or s2=="positive"):
        return ("positive")
    else:
        return ("non_digital")
    
    
if __name__ == '__main__':
    digital_positive=[]
    non_digital_positive=[]
    nond_negative=[]
    void=[]
    file_path = "C:\\Code_Sketch\\NPS\\S3134076\\PycharmProjects\\nps\\Data Set\\"
    url_path = 'http://localhost:8983/solr/'
    table_name='nps_Pipeline'
    solr_path = "C:\\Users\\s1761548\\Downloads\\solr-6.5.1\\solr-6.5.1\\bin"
    reader = xlrd.open_workbook(file_path + 'Scotiabank_Responses_Sept.xlsx')
    stat=dict((el,0) for el in json.loads(open('C:\\Code_Sketch\\NPS\\S3134076\\PycharmProjects\\nps\\Key_word.json').read())['cata_list'])
    stat_1=dict((el,0) for el in json.loads(open('C:\\Code_Sketch\\NPS\\S3134076\\PycharmProjects\\nps\\Nond_Key_Word.json').read())['cata_list'])
    stat = {**stat, **stat_1}
    stat['Navigation']=0
    stat['Non_Digital_Others']=0
    sh = reader.sheet_by_index(0)

    nps_data_pull()

    nps_prediction()

    data_writer()






